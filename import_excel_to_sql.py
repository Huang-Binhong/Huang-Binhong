#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
黄宾虹数据导入脚本
将 Excel 文件中的作品和事件数据转换为 SQL INSERT 语句
"""

import openpyxl
import json
import re
from datetime import datetime


def clean_text(text):
    """清理文本，移除多余空格和换行"""
    if text is None:
        return None
    return str(text).strip().replace('\n', ' ').replace('\r', '')


def extract_year_from_date(date_str):
    """从日期字符串中提取年份"""
    if not date_str:
        return None

    # 匹配各种年份格式
    patterns = [
        r'公元\s*(\d{4})\s*年',  # 公元 2009 年
        r'(\d{4})\s*年',          # 2009 年
        r'^(\d{4})$',             # 2009
    ]

    for pattern in patterns:
        match = re.search(pattern, str(date_str))
        if match:
            year = match.group(1)
            return f"{year}-01-01"

    return None


def extract_material(title):
    """从作品标题中提取材质"""
    if not title:
        return None

    materials = [
        '水墨纸本', '设色纸本', '墨笔纸本', '墨色纸本',
        '纸本', '绢本', '水墨绢本', '设色绢本'
    ]

    for material in materials:
        if material in title:
            return material

    return None


def parse_historical_events(events_str):
    """解析历史事件字符串为 JSON 数组"""
    if not events_str or events_str == '无明确对应历史事件':
        return '[]'

    # 分割事件（按数字序号分割）
    events = re.split(r'\d+\.\s*', str(events_str))
    events = [e.strip() for e in events if e.strip()]

    # 进一步分割（按分号或中文分号）
    result = []
    for event in events:
        sub_events = re.split(r'[;；]', event)
        result.extend([e.strip() for e in sub_events if e.strip()])

    return json.dumps(result, ensure_ascii=False)


def escape_sql_string(text):
    """转义 SQL 字符串"""
    if text is None:
        return 'NULL'
    return "'" + str(text).replace("'", "''") + "'"


def generate_works_sql(excel_file, output_file):
    """生成作品数据的 SQL INSERT 语句"""
    print(f"读取作品数据: {excel_file}")

    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active

    sql_statements = []
    sql_statements.append("-- 黄宾虹书法作品数据")
    sql_statements.append("-- 共 332 项作品\n")

    # 跳过标题行，从第2行开始
    for row_idx in range(2, sheet.max_row + 1):
        row = sheet[row_idx]

        # 读取各列数据
        title = clean_text(row[0].value)  # 作品标题
        period = clean_text(row[1].value)  # 历史年代
        creation_year = clean_text(row[2].value)  # 创作年代
        dimensions = clean_text(row[3].value)  # 尺寸
        seal = clean_text(row[4].value)  # 钤印
        inscription = clean_text(row[5].value)  # 款识
        image_url = clean_text(row[6].value)  # 图片URL（网址）
        # image_filename = clean_text(row[7].value)  # 图片文件名（不使用）

        if not title:
            continue

        # 提取创作年份（标准化）
        creation_date = extract_year_from_date(creation_year)

        # 提取材质（从标题中）
        material = extract_material(title)

        # 生成 SQL INSERT 语句（包含所有独立字段）
        sql = f"""INSERT INTO works (person_id, title, category, style_period, creation_year, dimensions, seal, inscription, material, creation_date, work_image_url) VALUES (
    1,
    {escape_sql_string(title)},
    '书法',
    {escape_sql_string(period)},
    {escape_sql_string(creation_year)},
    {escape_sql_string(dimensions)},
    {escape_sql_string(seal)},
    {escape_sql_string(inscription)},
    {escape_sql_string(material)},
    {escape_sql_string(creation_date)},
    {escape_sql_string(image_url)}
);"""

        sql_statements.append(sql)

    # 写入文件
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write('\n'.join(sql_statements))

    print(f"✅ 生成作品 SQL: {output_file}")
    print(f"   共 {len(sql_statements) - 2} 条记录")


def generate_events_sql(excel_file, output_file):
    """生成事件数据的 SQL INSERT 语句"""
    print(f"\n读取事件数据: {excel_file}")

    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active

    sql_statements = []
    sql_statements.append("-- 黄宾虹年份事件与历史事件")
    sql_statements.append("-- 生平事件数据\n")

    # 跳过标题行，从第2行开始
    for row_idx in range(2, sheet.max_row + 1):
        row = sheet[row_idx]

        # 读取各列数据
        year = clean_text(row[0].value)  # 年份
        event_text = clean_text(row[1].value)  # 事件
        development_phase = clean_text(row[2].value)  # 总体发展
        historical_events = clean_text(row[3].value)  # 历史事件

        if not year or not event_text:
            continue

        # 格式化日期
        event_date = f"{year}-01-01" if year.isdigit() else None

        # 解析历史事件为 JSON
        historical_events_json = parse_historical_events(historical_events)

        # 生成 SQL INSERT 语句
        sql = f"""INSERT INTO events (person_id, event_date, title, description, type, period, historical_events) VALUES (
    1,
    {escape_sql_string(event_date)},
    {escape_sql_string(event_text)},
    {escape_sql_string(event_text)},
    {escape_sql_string(development_phase)},
    {escape_sql_string(development_phase)},
    {escape_sql_string(historical_events_json)}
);"""

        sql_statements.append(sql)

    # 写入文件
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write('\n'.join(sql_statements))

    print(f"✅ 生成事件 SQL: {output_file}")
    print(f"   共 {len(sql_statements) - 2} 条记录")


def main():
    """主函数"""
    print("="*60)
    print("黄宾虹数据导入脚本")
    print("="*60)

    # 生成作品 SQL
    generate_works_sql(
        '332项黄宾虹书法作品.xlsx',
        'database/seed_huangbinhong_works.sql'
    )

    # 生成事件 SQL
    generate_events_sql(
        '黄宾虹年份事件与历史事件.xlsx',
        'database/seed_huangbinhong_events.sql'
    )

    print("\n" + "="*60)
    print("✅ 所有 SQL 文件生成完成！")
    print("="*60)
    print("\n下一步：")
    print("1. 检查生成的 SQL 文件")
    print("2. 执行导入：修改 database/db.go 添加这两个 seed 文件")
    print("3. 或手动执行：sqlite3 data/myapp.db < database/seed_huangbinhong_works.sql")


if __name__ == '__main__':
    main()
